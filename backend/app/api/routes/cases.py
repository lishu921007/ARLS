from datetime import date, datetime
from io import BytesIO, StringIO
import csv
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.db.session import get_db
from app.models.models import CaseRecord
from app.schemas.case import CaseCreate, CaseUpdate
from app.services.case_service import compute_case_fields, enqueue_warning_if_needed, priority_rank
from app.services.audit_service import write_audit

router = APIRouter(prefix='/cases', tags=['cases'])


def _apply_filters(query, keyword=None, notice_no=None, applicant=None, handling_department=None, case_type=None, decision_content=None, current_status=None, warning_status=None, closed_status=None, received_start=None, received_end=None, deadline_start=None, deadline_end=None, decision_start=None, decision_end=None):
    if keyword:
        query = query.filter((CaseRecord.notice_no.contains(keyword)) | (CaseRecord.applicant.contains(keyword)) | (CaseRecord.review_item.contains(keyword)) | (CaseRecord.respondent_subject.contains(keyword)) | (CaseRecord.review_content.contains(keyword)))
    if notice_no:
        query = query.filter(CaseRecord.notice_no.contains(notice_no))
    if applicant:
        query = query.filter(CaseRecord.applicant.contains(applicant))
    if handling_department:
        query = query.filter(CaseRecord.handling_department == handling_department)
    if case_type:
        query = query.filter(CaseRecord.case_type == case_type)
    if decision_content:
        query = query.filter(CaseRecord.decision_content == decision_content)
    if current_status:
        query = query.filter(CaseRecord.current_status == current_status)
    if warning_status:
        query = query.filter(CaseRecord.warning_status == warning_status)
    if closed_status == 'closed':
        query = query.filter(CaseRecord.is_closed == True)
    elif closed_status == 'open':
        query = query.filter(CaseRecord.is_closed == False)
    if received_start:
        query = query.filter(CaseRecord.received_date >= received_start)
    if received_end:
        query = query.filter(CaseRecord.received_date <= received_end)
    if deadline_start:
        query = query.filter(CaseRecord.deadline_date >= deadline_start)
    if deadline_end:
        query = query.filter(CaseRecord.deadline_date <= deadline_end)
    if decision_start:
        query = query.filter(CaseRecord.decision_date >= decision_start)
    if decision_end:
        query = query.filter(CaseRecord.decision_date <= decision_end)
    return query


def _resequence_cases(db: Session):
    rows = db.query(CaseRecord).filter(CaseRecord.deleted == False).order_by(CaseRecord.created_at.asc(), CaseRecord.id.asc()).all()
    changed = False
    for idx, row in enumerate(rows, start=1):
        if row.seq_no != idx:
            row.seq_no = idx
            changed = True
    if changed:
        db.flush()


def _load_rows(db: Session, **filters):
    query = db.query(CaseRecord).filter(CaseRecord.deleted == False)
    query = _apply_filters(query, **filters)
    rows = query.all()
    for row in rows:
        compute_case_fields(db, row)
    _resequence_cases(db)
    db.commit()
    rows.sort(key=lambda x: (priority_rank(x), x.remaining_workdays if x.remaining_workdays is not None else 9999, x.deadline_date or date.max, x.created_at))
    return rows


@router.get('')
def list_cases(
    keyword: str | None = None,
    notice_no: str | None = None,
    applicant: str | None = None,
    handling_department: str | None = None,
    case_type: str | None = None,
    decision_content: str | None = None,
    current_status: str | None = None,
    warning_status: str | None = None,
    closed_status: str | None = None,
    received_start: date | None = None,
    received_end: date | None = None,
    deadline_start: date | None = None,
    deadline_end: date | None = None,
    decision_start: date | None = None,
    decision_end: date | None = None,
    db: Session = Depends(get_db),
):
    return _load_rows(db,
        keyword=keyword, notice_no=notice_no, applicant=applicant,
        handling_department=handling_department, case_type=case_type,
        decision_content=decision_content, current_status=current_status,
        warning_status=warning_status, closed_status=closed_status,
        received_start=received_start, received_end=received_end,
        deadline_start=deadline_start, deadline_end=deadline_end,
        decision_start=decision_start, decision_end=decision_end,
    )


@router.get('/export')
def export_cases(
    format: str = Query('csv'),
    keyword: str | None = None,
    notice_no: str | None = None,
    applicant: str | None = None,
    handling_department: str | None = None,
    case_type: str | None = None,
    decision_content: str | None = None,
    current_status: str | None = None,
    warning_status: str | None = None,
    closed_status: str | None = None,
    received_start: date | None = None,
    received_end: date | None = None,
    deadline_start: date | None = None,
    deadline_end: date | None = None,
    decision_start: date | None = None,
    decision_end: date | None = None,
    db: Session = Depends(get_db),
):
    rows = _load_rows(db,
        keyword=keyword, notice_no=notice_no, applicant=applicant,
        handling_department=handling_department, case_type=case_type,
        decision_content=decision_content, current_status=current_status,
        warning_status=warning_status, closed_status=closed_status,
        received_start=received_start, received_end=received_end,
        deadline_start=deadline_start, deadline_end=deadline_end,
        decision_start=decision_start, decision_end=decision_end,
    )
    headers = ['序号', '通知书编号', '申请人', '涉案主体', '行政复议事项', '复议内容', '经办部门', '联系人', '签收日期', '最晚答复日', '实际答复时间', '决定时间', '决定内容', '类型', '当前状态', '是否闭环', '预警状态', '剩余工作日', '备注']
    if format == 'csv':
        sio = StringIO(); writer = csv.writer(sio); writer.writerow(headers)
        for x in rows:
            writer.writerow([x.seq_no, x.notice_no, x.applicant, x.respondent_subject or '', x.review_item, x.review_content or '', x.handling_department, x.contact_name or '', x.received_date or '', x.deadline_date or '', x.actual_reply_time or '', x.decision_date or '', x.decision_content or '', x.case_type or '', x.current_status, '已闭环' if x.is_closed else '未闭环', x.warning_status, x.remaining_workdays if x.remaining_workdays is not None else '', x.remark or ''])
        output = BytesIO(sio.getvalue().encode('utf-8-sig'))
        write_audit(db, 'case.export.csv', 'case', detail=f'count={len(rows)}')
        return StreamingResponse(output, media_type='text/csv', headers={'Content-Disposition': f"attachment; filename=cases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"})
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = '事项台账'; ws.append(headers)
    for x in rows:
        ws.append([x.seq_no, x.notice_no, x.applicant, x.respondent_subject or '', x.review_item, x.review_content or '', x.handling_department, x.contact_name or '', str(x.received_date or ''), str(x.deadline_date or ''), str(x.actual_reply_time or ''), str(x.decision_date or ''), x.decision_content or '', x.case_type or '', x.current_status, '已闭环' if x.is_closed else '未闭环', x.warning_status, x.remaining_workdays if x.remaining_workdays is not None else '', x.remark or ''])
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    write_audit(db, 'case.export.excel', 'case', detail=f'count={len(rows)}')
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f"attachment; filename=cases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"})


@router.get('/import-template')
def download_import_template(format: str = 'xlsx'):
    headers = ['通知书编号', '申请人', '涉案主体', '行政复议事项', '复议内容', '经办部门', '联系人', '签收日期', '决定内容', '类型', '当前状态', '备注']
    sample = ['2026复议001', '张三', '某公司', '行政处罚', '不服行政处罚决定，申请行政复议', '浆洗所', '李四', '2026-04-21', '维持', '食品', '待处理', '示例数据']
    if format == 'csv':
        sio = StringIO(); writer = csv.writer(sio); writer.writerow(headers); writer.writerow(sample)
        return StreamingResponse(BytesIO(sio.getvalue().encode('utf-8-sig')), media_type='text/csv', headers={'Content-Disposition': 'attachment; filename=case_import_template.csv'})
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = '导入模板'; ws.append(headers); ws.append(sample)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=case_import_template.xlsx'})


@router.post('/import')
def import_cases(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = file.file.read()
    imported = 0
    errors = []
    next_seq = (db.query(CaseRecord.seq_no).filter(CaseRecord.deleted == False).order_by(CaseRecord.seq_no.desc(), CaseRecord.id.desc()).first() or [0])[0] or 0

    def add_case_from_row(row, line_no):
        nonlocal imported, next_seq
        try:
            next_seq += 1
            case = CaseRecord(
                seq_no=next_seq,
                notice_no=(row.get('通知书编号', '') or '').strip() if isinstance(row.get('通知书编号', ''), str) else (row.get('通知书编号', '') or ''),
                applicant=(row.get('申请人', '') or '').strip() if isinstance(row.get('申请人', ''), str) else (row.get('申请人', '') or ''),
                respondent_subject=row.get('涉案主体', '') or '',
                review_item=row.get('行政复议事项', '') or '',
                review_content=row.get('复议内容', '') or '',
                handling_department=row.get('经办部门', '') or '',
                contact_name=row.get('联系人', '') or '',
                current_status=row.get('当前状态', '待处理') or '待处理',
                decision_content=row.get('决定内容', '') or '',
                case_type=row.get('类型', '') or '',
                remark=row.get('备注', '') or '',
            )
            rv = row.get('签收日期')
            if rv:
                try:
                    case.received_date = rv if isinstance(rv, date) else datetime.strptime(str(rv), '%Y-%m-%d').date()
                except Exception:
                    next_seq -= 1
                    errors.append({'line': line_no, 'reason': f'签收日期格式错误: {rv}'})
                    return
            compute_case_fields(db, case)
            db.add(case)
            imported += 1
        except Exception as e:
            next_seq -= 1
            errors.append({'line': line_no, 'reason': str(e)})

    if file.filename.lower().endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(StringIO(text))
        for i, row in enumerate(reader, start=2):
            add_case_from_row(row, i)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for i, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = dict(zip(headers, values))
            add_case_from_row(row, i)

    _resequence_cases(db)
    db.commit()
    write_audit(db, 'case.import', 'case', detail=f'count={imported}, errors={len(errors)}')
    return {'success': True, 'imported': imported, 'errors': errors}


@router.post('')
def create_case(payload: CaseCreate, db: Session = Depends(get_db)):
    seq = (db.query(CaseRecord.seq_no).filter(CaseRecord.deleted == False).order_by(CaseRecord.seq_no.desc(), CaseRecord.id.desc()).first() or [0])[0] or 0
    row = CaseRecord(seq_no=seq + 1, **payload.model_dump())
    if not row.current_status:
        row.current_status = '待处理'
    compute_case_fields(db, row)
    db.add(row)
    _resequence_cases(db)
    db.commit()
    db.refresh(row)
    enqueue_warning_if_needed(db, row)
    write_audit(db, 'case.create', 'case', str(row.id), detail=row.notice_no)
    return row


@router.get('/{case_id}')
def get_case(case_id: int, db: Session = Depends(get_db)):
    row = db.query(CaseRecord).filter(CaseRecord.id == case_id, CaseRecord.deleted == False).first()
    if not row:
        raise HTTPException(404, '事项不存在')
    compute_case_fields(db, row)
    db.flush()
    db.refresh(row)
    return row


@router.put('/{case_id}')
def update_case(case_id: int, payload: CaseUpdate, db: Session = Depends(get_db)):
    row = db.query(CaseRecord).filter(CaseRecord.id == case_id, CaseRecord.deleted == False).first()
    if not row:
        raise HTTPException(404, '事项不存在')
    for k, v in payload.model_dump().items():
        setattr(row, k, v)
    if not row.current_status:
        row.current_status = '待处理'
    compute_case_fields(db, row)
    db.commit(); db.refresh(row)
    write_audit(db, 'case.update', 'case', str(row.id), detail=row.notice_no)
    return row


@router.post('/{case_id}/close')
def close_case(case_id: int, db: Session = Depends(get_db)):
    row = db.query(CaseRecord).filter(CaseRecord.id == case_id, CaseRecord.deleted == False).first()
    if not row:
        raise HTTPException(404, '事项不存在')
    row.is_closed = True; row.current_status = '已办结'; compute_case_fields(db, row); db.commit()
    write_audit(db, 'case.close', 'case', str(row.id), detail=row.notice_no)
    return {'success': True}


@router.delete('/{case_id}')
def delete_case(case_id: int, db: Session = Depends(get_db)):
    row = db.query(CaseRecord).filter(CaseRecord.id == case_id).first()
    if not row:
        raise HTTPException(404, '事项不存在')
    row.deleted = True
    _resequence_cases(db)
    db.commit()
    write_audit(db, 'case.delete', 'case', str(row.id), detail=row.notice_no)
    return {'success': True}
